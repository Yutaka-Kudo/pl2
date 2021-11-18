from site_package.my_module import ZEN2HAN, pl_dbdict

from pprint import pprint as pp
import openpyxl
from django.http.response import HttpResponse
from django.shortcuts import redirect, render
# import os


def index(request):
    return render(request, "upload_part/upload.html")


def admin_scr(request):
    return redirect('/admin/')


# UPLOAD_DIR = os.path.dirname(os.path.abspath(__file__)) + "/media/"
wb = ""


def show_sheetname(request):
    global wb
    wb = openpyxl.load_workbook(request.FILES['testfile'])
    print(wb.sheetnames)
    month_list = [str(month) for month in range(1, 13)]
    return render(request, "upload_part/choice_sheet.html", {"sheet_list": wb.sheetnames, "month_list": month_list})


def excel_up(request):
    sheet_name = request.GET.get('sheet')
    month = request.GET.get('month')

    # wb = openpyxl.load_workbook(request.FILES['testfile'].file)
    # wb.sheetnames
    sheet = wb[sheet_name]

    # # まとめて取得用 
    # months = month
    # for month in range(int(months),13):
    #     month = str(month)


    counter = 0
    # シート内最終カラム番号取得
    r, c = 4, 2
    while sheet.cell(r, c).value is not None:
        c += 1
    last_c = c - 1
    print(last_c)

    # 最終ロウ番号取得
    r = 6
    while sheet.cell(r, 1).value is not None or sheet.cell(r, 3).value is not None or sheet.cell(r+1, 1).value is not None:
        r += 1
    print(r)
    last_r = r

    # 開始カラムと終了カラム
    for i in range(2, last_c+1):
        if sheet.cell(2, i).value and len(sheet.cell(2, i).value) <= 4:
            if sheet.cell(2, i).value[:-1].translate(ZEN2HAN) == month:
                st_C = i
                for ii in range(st_C+1, last_c+1):
                    if sheet.cell(2, ii).value is not None:
                        end_col = ii - 1
                        break
                    if ii == last_c:
                        end_col = ii
                        break
    st_R = 2
    month_R, month_C = 2, 2
    print(st_C, end_col)

    message_list = []

    for col in range(st_C, end_col+1):
        # if sheet.cell(st_R, st_C).value is not None:  # 月
        if sheet.cell(st_R+2, col).value == "売上高":
            value_list = []
            pay_type_list = ["日付", "カテゴリー"]

            # year,month
            # year = "2020" # 12月が次年度のシートにある場合。数字を適宜変更
            year = sheet["a2"].value[:-1]
            month_C = st_C
            month = sheet.cell(month_R, month_C).value[:-1].translate(ZEN2HAN).zfill(2)
            y_m = year+month
            value_list.append(y_m)

            # category
            category = sheet.cell(st_R+1, col).value
            value_list.append(category)

            # 完了画面のメッセージ用
            totalsales_for_message = sheet.cell(st_R+3, col).value
            message_list.extend([category, totalsales_for_message])

            i = 5
            while i < last_r:
                pay_value = sheet.cell(i, col).value
                pay_type = sheet[f"a{i}"].value
                if pay_value is None:
                    pay_value = 0
                if type(pay_value) == int and pay_type is not None:
                    print(pay_type, pay_value)
                    value_list.append(pay_value)
                    pay_type_list.append(pay_type)
                i += 1

            # 1つずつ既存のモデルフィールドと照会し、removeしていく。1つでも余ればエラーを返す。
            ver_name_list = []
            for each_dict in pl_dbdict.values():
                for ver_name in each_dict.values():
                    ver_name_list.append(ver_name)
            compare_list = [s for s in pay_type_list if s not in ver_name_list]
            if compare_list:
                # raise Exception(f'項目名が一致しません{compare_list}')
                return HttpResponse(f'項目名が一致しません{compare_list}')


            exceldata = dict(zip(pay_type_list[2:], value_list[2:]))  # 最初の2個は年月と店名なのでスキップ
            for pay_type, value in exceldata.items():
                for model, each_dict in pl_dbdict.items():
                    for db_column_name, ver_name in each_dict.items():
                        if pay_type == ver_name:
                            if model.__name__ == 'PL_data':
                                pl_model, _ = model.objects.update_or_create(
                                    y_m=value_list[0], category=value_list[1],
                                    defaults={
                                        db_column_name: value,
                                    })
                            else:
                                model.objects.update_or_create(
                                    pl_data=pl_model,
                                    defaults={
                                        db_column_name: value,
                                    })
            counter += 1

            print(counter)

    return render(request, 'upload_part/upload.html', {"messages": message_list})


# ----sql------
# config = {
#     "user": "yutakakudo",
#     "password": "udondondon",
#     "host": "localhost",
#     "database": "syuushihyou",
# }
# cnx = mysql.connector.connect(**config)
# cur = cnx.cursor()
# cur.execute('desc upload_part_PL_data')
# rows = cur.fetchall()
# for row in rows:
#     print(row[0])
# cnx.close()
# ----sql------
