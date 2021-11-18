from icecream import ic
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
import django_pandas.io as pd
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl.styles import PatternFill, borders
import openpyxl
import pandas

from decimal import Decimal, ROUND_HALF_UP
import datetime as dt
from dateutil.relativedelta import relativedelta
import urllib.parse
from io import BytesIO
import urllib

from config import settings
from site_package import my_module
from upload_part.models import PL_data, FoodCosts, DrinkCosts, LaborCosts, UtilityCosts_ComunicationCosts, AdvertisingCosts, OtherCosts, TaxExemptExpenses
from prediction.models import PRED_data
from daily_report.actions import graph_helper
from daily_report import models as drmodels
from .helpers import auth_helper
from . import models

from dotenv import load_dotenv
load_dotenv()

ic.configureOutput(prefix='', includeContext=True)
# ic.disable()
# set_level = '' # 全てに影響
set_level = 'debug'
# set_level = 'info'
logger = my_module.create_logger(__name__, set_level=set_level)

# import matplotlib.pyplot as plt
# import japanize_matplotlib
# import matplotlib
# import matplotlib.style as mplstyle
# import matplotlib.collections as mc
# matplotlib.use('Agg')


qall = PL_data.objects.all().order_by("-y_m", "id")
category_list = []
for i in qall:
    category_list.append(i.category)
    if i == qall[0]:
        ym = i.y_m
category_list = sorted(set(category_list), key=category_list.index)  # 他からimportしてる

today = dt.datetime.now().date()
pred_query = PRED_data.objects.filter(date__gte=today).order_by('date')
pred_date_list = [str(q.date) for q in pred_query]
pred_date_list_reversed = list(reversed(pred_date_list))  # 他からimportしてる

previous_query = PRED_data.objects.filter(date__lte=today).order_by('date')
previous_date_list = list(reversed(sorted(set([str(q.date)[:7] for q in previous_query]))))  # 他からimportしてる


def pl_ini():
    now_year = dt.datetime.now().year
    year_list = []
    first_year = 2017
    while now_year >= first_year:
        year_list.append(now_year)
        now_year -= 1
    month_list = [i for i in reversed(range(1, 13))]

    year_list_fes = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="FES")]), reverse=True)
    year_list_garage = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="Garage あそび")]), reverse=True)
    year_list_wanaichi = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="罠一目")]), reverse=True)
    year_list_wananakame = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="罠中目黒")]), reverse=True)
    year_list_tourou = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="灯篭")]), reverse=True)
    year_list_baseasobi = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="Base\u3000あそび")]), reverse=True)
    year_list_urayasu = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="浦安あそび")]), reverse=True)
    year_list_honnbu = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="本部")]), reverse=True)
    year_list_storeTotal = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="店舗合計")]), reverse=True)
    year_list_total = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="合計")]), reverse=True)
    year_list_reme = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="reme(小顔矯正)")]), reverse=True)
    year_list_tanoshi = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="楽")]), reverse=True)
    year_list_wana = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="罠")]), reverse=True)
    year_list_yotteko = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="よってこ")]), reverse=True)
    year_list_yottekoEntrust = sorted(set([q.y_m[:-2] for q in PL_data.objects.filter(category="よってこ業務委託料")]), reverse=True)

    context = {
        "cgy_list": category_list,
        "year_list": year_list,
        "month_list": month_list,

        "year_list_fes": year_list_fes,
        "year_list_garage": year_list_garage,
        "year_list_wanaichi": year_list_wanaichi,
        "year_list_wananakame": year_list_wananakame,
        "year_list_tourou": year_list_tourou,
        "year_list_baseasobi": year_list_baseasobi,
        "year_list_urayasu": year_list_urayasu,
        "year_list_honnbu": year_list_honnbu,
        "year_list_storeTotal": year_list_storeTotal,
        "year_list_total": year_list_total,
        "year_list_reme": year_list_reme,
        "year_list_tanoshi": year_list_tanoshi,
        "year_list_wana": year_list_wana,
        "year_list_yotteko": year_list_yotteko,
        "year_list_yottekoEntrust": year_list_yottekoEntrust
    }
    return context


def index(request):
    context = initialize_context(request)

    store_dict = {"FES": "fes", "Garage": "garage", "灯篭": "tourou", "罠一目": "wanaichi", "罠中目黒": "wananakame"}

    context.update({
        "pred_date_list": pred_date_list,
        "pred_date_list_reversed": pred_date_list_reversed,
        "previous_date_list": previous_date_list,
        "store_dict": store_dict,
    })

    if context['user'].get('email'):
        _, store_name = graph_helper.match_stores_by_email(context['user']['email'])
        context['store_name'] = store_name

        confirm_permission(request)
        context['user']['super'] = request.session['user'].get('super')
        context['user']['manager'] = request.session['user'].get('manager')

        if request.session['user'].get('manager'):
            context.update(pl_ini())

    if settings.DEBUG:
        request.session['debug'] = 'true'
    else:
        request.session['debug'] = 'false'

    ic(context['user'])
    return render(request, "display_part/index.html", context)


def confirm_permission(request):
    """sessionを変更"""
    mng_objs = models.Managers.objects.all()
    if request.session['user']['email'] in [v['email'] for v in mng_objs.filter(c_d_permission=True).values()]:
        request.session['user']['super'] = True
    else:
        request.session['user']['super'] = False

    if request.session['user']['email'] in [v['email'] for v in mng_objs.values()]:
        request.session['user']['manager'] = True
    else:
        request.session['user']['manager'] = False


def initialize_context(request):
    context = {}

    # Check for any errors in the session
    error = request.session.pop('flash_error', None)

    if error is not None:
        context['errors'] = []
        context['errors'].append(error)

    # Check for user in the session
    context['user'] = request.session.get('user', {'is_authenticated': False})
    return context


def sign_in(request):
    # Get the sign-in flow
    flow = auth_helper.get_sign_in_flow()
    # Save the expected flow so we can use it in the callback
    try:
        request.session['auth_flow'] = flow
    except Exception as e:
        logger.error(e)
    # Redirect to the Azure sign-in page

    ic('sign in')
    return HttpResponseRedirect(flow['auth_uri'])


def callback(request):
    # Make the token request
    result = auth_helper.get_token_from_code(request)

    # Get the user's profile
    user = graph_helper.get_user(result['access_token'])

    ic(user)

    # Store user
    auth_helper.store_user(request, user)

    # return HttpResponseRedirect(reverse('daily_report:home'))
    return HttpResponseRedirect(reverse('display_part:index'))


def sign_out(request):
    # Clear out the user and token
    auth_helper.remove_user_and_token(request)

    return HttpResponseRedirect(reverse('display_part:index'))


def manage(request):
    confirm_permission(request)

    dbdict = my_module.create_fieldname_dict(drmodels.Costs)
    costs = list(drmodels.Costs.objects.all().order_by('id').values())

    managers = list(models.Managers.objects.all().order_by('id').values('id', 'name', 'email', 'c_d_permission'))

    now = dt.datetime.now()
    month_list = []
    for i in range(1, 6):
        to_month = now - relativedelta(months=i)
        month_list.append(dt.datetime.strftime(to_month, '%Y-%m'))

    context = {
        'dbdict': dbdict,
        'costs': costs,
        'managers': managers,
        'user': request.session.get('user', None),
        'month_list': month_list,
    }
    # ic(context)

    return render(request, 'display_part/manage.html', context)


def manage_costs(request):
    request_gets: dict = request.GET.dict()
    id = request_gets.pop('id')
    ic(request_gets)

    if request_gets.pop('update', None):
        drmodels.Costs.objects.filter(id=id).update(**request_gets)

    is_auto = [k for k in request_gets if 'auto_' in k]
    if is_auto:
        request_gets.pop(is_auto[0], None)

        def get_average_costs(store: str, year: str, selected_kinds: str, drop0: bool = False):
            df = create_df_by_paytype(year, selected_kinds)
            values = df.loc[store].drop(['店名', '合計']).tolist()
            if drop0:
                values = [v for v in values if v]
            average = sum(values) / len(values)
            ic(average)
            return average

        costname_in_models = is_auto[0].replace('auto_', '')
        store = drmodels.Costs.objects.get(id=id).store
        store = my_module.translate_storename(store, to_ja=True)
        year = str(dt.datetime.now().year)
        kinds_list, minus_list = [], []
        if costname_in_models == 'ad_cost':
            kinds_list = ['宣伝・広告費　合計']
            minus_list = []
        elif costname_in_models == 'utility_cost':
            kinds_list = ['ガス代', '電気代', '上水道代', '下水道']
            minus_list = []
        elif costname_in_models == 'other_total':
            kinds_list = ['経費　総計']
            minus_list = ['宣伝・広告費　合計', 'ガス代', '電気代', '上水道代', '下水道', '家賃', '家賃　更新代']

        total = 0
        for kinds in kinds_list:
            total += get_average_costs(store, year, kinds)
        minus = 0
        for kinds in minus_list:
            minus += get_average_costs(store, year, kinds)
        drmodels.Costs.objects.filter(id=id).update(**{costname_in_models: total - minus})
    return HttpResponseRedirect(reverse('display_part:manage'))


def manage_manager(request):
    id = request.GET.get('id')
    if request.GET.get('delete'):
        models.Managers.objects.get(id=id).delete()
        return HttpResponseRedirect(reverse('display_part:manage'))

    name = request.GET.get('name')
    email = request.GET.get('email')
    c_d_permission = request.GET.get('c_d_permission')
    if request.GET.get('update'):
        models.Managers.objects.filter(id=id).update(
            name=name,
            email=email,
            c_d_permission=True if c_d_permission else False,
        )
    if request.GET.get('regist'):
        models.Managers.objects.create(
            name=name,
            email=email,
            c_d_permission=True if c_d_permission else False
        )

    ic(name, email, c_d_permission)
    # if c_d_permission:

    return HttpResponseRedirect(reverse('display_part:manage'))


def detailbymonth(request, when=""):
    if when and when != "xxx":
        when = urllib.parse.unquote(when).replace("/", "")
    else:
        y1 = request.GET.get('year-select')
        m1 = request.GET.get('month-select').zfill(2)
        # y2 = request.GET.get('year-select2')
        # m2 = request.GET.get('month-select2').zfill(2)
        when = y1 + m1
        # when2 = y2 + m2

    print(when)
    # print(when2)

    q_list = [when]

    # # どちらかのフォームが未選択の場合のフォロー
    # q_list = []
    # if "default" in when1 or when1 == "":
    #     q_list = [when2]
    # elif "default" in when2 or when2 == "":
    #     q_list = [when1]
    # else:
    #     when1, when2 = dt.datetime.strptime(when1, "%Y%m"), dt.datetime.strptime(when2, "%Y%m")
    #     if (when1) < (when2):
    #         while (when1) <= (when2):
    #             q_list.append(str(when1.year)+str(when1.month).zfill(2))
    #             when1 += relativedelta(months=1)
    #     else:
    #         while (when1) >= (when2):
    #             q_list.append(str(when2.year)+str(when2.month).zfill(2))
    #             when2 += relativedelta(months=1)
    # # -------------------------------------
    request.session["q_list"] = q_list
    request.session["when_session"] = when[:-2]

    df = create_df_month(q_list, use="table")
    if df is None:
        return HttpResponse(f'{q_list} 指定された期間が存在しません。')

    context = {
        # "data_value": pl_data,
        "df": df,
    }
    return render(request, 'display_part/detail.html', context)


def detailbystore(request, store="", when=""):
    if store != "xxx":
        store = urllib.parse.unquote(store)
        when = urllib.parse.unquote(when)
        when = when.split("/")[0]
    else:
        store = request.GET.get('store')
        when = request.GET.get('year')

    request.session["cgy_session"] = store
    request.session["when_session"] = when

    df2, pl_data = create_df_by_store(store, when, use="table")

    context = {
        "data_value": pl_data,
        "df2": df2,
        "cgy": store,
        # "when": when,

    }
    return render(request, 'display_part/detailbystore.html', context)


def detailbypaytype(request, kinds=""):
    kinds = urllib.parse.unquote(kinds)
    year = request.session["when_session"].split("/")[0]

    df = create_df_by_paytype(year, kinds)

    context = {
        "df": df,
        "kinds": kinds,
        "year": year,
    }
    return render(request, 'display_part/table_by_paytype.html', context)


def output_excel(request, flag):
    if flag == "1":
        q_list = request.session["q_list"]
        df_for_excel = create_df_month(q_list, use="excel")

        q = q_list[0]+"~"+q_list[-1]

        filename = '収支表{}.xlsx'.format(q)
        quoted_filename = urllib.parse.quote(filename)

        bio = BytesIO()
        writer = pandas.ExcelWriter(bio, engine="xlsxwriter")
        df_for_excel.set_index('日付', inplace=True)
        df_for_excel.T.to_excel(writer, float_format='%.2f')
        writer.save()
        bio.seek(0)

    elif flag == "2":
        cgy = request.session["cgy_session"]
        when = request.session["when_session"]
        df_for_excel2 = create_df_by_store(cgy, when, use="excel")

        filename = f'収支表_{cgy}_{when}.xlsx'
        quoted_filename = urllib.parse.quote(filename)

        bio = BytesIO()
        writer = pandas.ExcelWriter(bio, engine="xlsxwriter")
        df_for_excel2.set_index('日付', inplace=True)
        df_for_excel2.T.to_excel(writer, float_format='%.2f')
        writer.save()
        bio.seek(0)

    # 数字をアルファベッドに変換！の関数
    def num2alpha(num):
        if num <= 26:
            return chr(64+num)
        elif num % 26 == 0:
            return num2alpha(num//26-1)+chr(90)
        else:
            return num2alpha(num//26)+chr(64+num % 26)
    # ----------------------------------

    wbk = openpyxl.load_workbook(bio)
    sheet = wbk["Sheet1"]

    for i in reversed(range(3, len(sheet[1])+1)):
        sheet.insert_cols(i)

    side = borders.Side(style="thin", color="000000")
    border = borders.Border(top=side, right=side, bottom=side, left=side)

    fill_ukon = PatternFill(patternType='solid', fgColor='fdb933')
    fill_skygray = PatternFill(patternType='solid', fgColor='d8d8d8')
    fill_gray = PatternFill(patternType='solid', fgColor='b1b3b6')

    fill_creamyellow = PatternFill(patternType='solid', fgColor='ffedb3')
    fill_usukihanada = PatternFill(patternType='solid', fgColor='adcdec')
    fill_paleorange = PatternFill(patternType='solid', fgColor='ffe6ce')
    # fill_paleorange = PatternFill(patternType='solid', fgColor='ffe6ce')

    sheet.cell(1, len(sheet[1])+1).value = ">"  # 下の処理中に長さが変わるため前処理。
    for row in sheet:
        for cell in row:
            try:
                if "-" in sheet.cell(1, cell.column).value:
                    sheet[cell.coordinate].fill = fill_paleorange
            except TypeError:
                pass
            if row[0].value == "総売上高":
                sheet[cell.coordinate].fill = fill_ukon
            if row[0].value == "フード合計":
                sheet[cell.coordinate].fill = fill_skygray
            if row[0].value == "ドリンク合計":
                sheet[cell.coordinate].fill = fill_skygray
            if row[0].value == "仕入合計":
                sheet[cell.coordinate].fill = fill_gray
            if row[0].value == "人件費合計":
                sheet[cell.coordinate].fill = fill_gray
            if row[0].value == "Ｆ/Ｌ コスト":
                sheet[cell.coordinate].fill = fill_gray
            if row[0].value == "営業粗利":
                sheet[cell.coordinate].fill = fill_creamyellow
            if row[0].value == "光熱費・通信費　合計":
                sheet[cell.coordinate].fill = fill_skygray
            if row[0].value == "宣伝・広告費　合計":
                sheet[cell.coordinate].fill = fill_skygray
            if row[0].value == "雑費　合計":
                sheet[cell.coordinate].fill = fill_skygray
            if row[0].value == "課税経費　合計":
                sheet[cell.coordinate].fill = fill_gray
            if row[0].value == "非課税経費　合計":
                sheet[cell.coordinate].fill = fill_gray
            if row[0].value == "経費　総計":
                sheet[cell.coordinate].fill = fill_gray
            if row[0].value == "経常利益額":
                sheet[cell.coordinate].fill = fill_creamyellow
            if row[0].value == "税引き後利益":
                sheet[cell.coordinate].fill = fill_usukihanada
            try:
                if cell.row >= 4:
                    if cell.column % 2 == 1 and cell.column != 1:
                        cell.value = f"={num2alpha(cell.column-1)}{cell.row}/{num2alpha(cell.column-1)}3"
                        cell.number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE_00

                        sheet.column_dimensions[num2alpha(cell.column)].width = 9
                    else:
                        sheet.column_dimensions[num2alpha(cell.column)].width = 12

            except TypeError:
                pass
            if sheet.cell(2, cell.column).value:
                cell.number_format = '#,##0;[red]-#,##0'
            sheet[cell.coordinate].border = border
    sheet.column_dimensions["A"].width = 40
    sheet.freeze_panes = "b4"

    # 自動アジャスト
    # for col in sheet.columns:
    # max_length = 0
    # for cell in col:
    #     if max_length < len(str(cell.value)):
    #         max_length = len(str(cell.value))
    # adjusted_width = (max_length + 1) * 1.2
    # sheet.column_dimensions[num2alpha(col[0].column)].width = 10

    # if "-" in sheet.cell(1, 2).value:
    #     sheet[sheet.cell(3, 2).coordinate].fill = fill
    #     sheet.cell(3, 2).number_format = openpyxl.styles.numbers.FORMAT_NUMBER
    #     sheet.cell(4, 2).number_format = openpyxl.styles.numbers.FORMAT_NUMBER_00
    #     sheet.cell(5, 2).number_format = openpyxl.styles.numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    #     sheet.cell(6, 2).number_format = openpyxl.styles.numbers.FORMAT_NUMBER_COMMA_SEPARATED2
    # sheet.cell(3, 2).number_format = '#,##0'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=UTF-8-sig')
    response['Content-Disposition'] = "attachment; filename={}; filename*=UTF-8-sig''{}".format(quoted_filename, quoted_filename)

    wbk.save(response)

    return response


def chart(request, flg):
    if flg == "month":
        q_list = request.session["q_list"]
        df_list_chart = create_df_month(q_list, use="chart")

        df = pandas.DataFrame()
        for df_for_chart in df_list_chart:
            if int(df_for_chart.loc[0, "日付"][:-3]) <= 2019:
                df_one = df_for_chart.loc[:, ["日付", "総売上高", "フード合計", "ドリンク合計", "人件費合計", "経費　総計", "経常利益額"]]
                for index in df_one.index:
                    # 経常利益額がマイナスなら、ゼロに。
                    if df_one.loc[index, "経常利益額"] < 0:
                        df_one.loc[index, "経常利益額"] = 0
            else:
                df_one = df_for_chart.loc[:, ["日付", "総売上高", "フード合計", "ドリンク合計", "社員人件費", "社保合計", "賞与", "アルバイト人件費", "賞与社保合計/その他（日払い等）", "交通費", "経費　総計", "経常利益額"]]

                df_one.insert(4, "社員", df_one.loc[:, "社員人件費"]+df_one.loc[:, "社保合計"]+df_one.loc[:, "賞与"])
                df_one.insert(5, "アルバイト", df_one.loc[:, "アルバイト人件費"]+df_one.loc[:, "賞与社保合計/その他（日払い等）"])

                for index in df_one.index:
                    # 経常利益額がマイナスなら、ゼロに。
                    if df_one.loc[index, "経常利益額"] < 0:
                        df_one.loc[index, "経常利益額"] = 0

                    # 交通費を半分ずつ分散。バイト人件費ゼロなら、全て社員に。
                    if df_one.loc[index, "アルバイト人件費"] == 0:
                        df_one.loc[index, "社員"] += df_one.loc[index, "交通費"]
                    else:
                        df_one.loc[index, "社員"] += round(df_one.loc[index, "交通費"] / 2)
                        df_one.loc[index, "アルバイト"] += round(df_one.loc[index, "交通費"] / 2)

                df_one.drop(columns=["社員人件費", "社保合計", "賞与", "アルバイト人件費", "賞与社保合計/その他（日払い等）", "交通費"], inplace=True)

                df = pandas.concat([df, df_one])

        df["総売上高"] = df["総売上高"].astype(int)
        # print(df)

        date_list = list(df["日付"])
        date_list = [date for date in list(df["日付"])]
        sales_list = list(df["総売上高"])
        # debug(date_list)

    elif flg == "cgy":
        cgy = request.session["cgy_session"]
        span = request.session["when_session"]
        df_for_chart = create_df_by_store(cgy, span, use="chart")

        if span == "2019" or span == "2018" or span == "2017":
            df = df_for_chart.loc[:, ["日付", "総売上高", "フード合計", "ドリンク合計", "人件費合計", "経費　総計", "経常利益額"]]
            for index in df.index:
                # 経常利益額がマイナスなら、ゼロに。
                if df.loc[index, "経常利益額"] < 0:
                    df.loc[index, "経常利益額"] = 0
        else:
            df = df_for_chart.loc[:, ["日付", "総売上高", "フード合計", "ドリンク合計", "社員人件費", "社保合計", "賞与", "アルバイト人件費", "賞与社保合計/その他（日払い等）", "交通費", "経費　総計", "経常利益額"]]

            df.insert(4, "社員", df.loc[:, "社員人件費"]+df.loc[:, "社保合計"]+df.loc[:, "賞与"])
            df.insert(5, "アルバイト", df.loc[:, "アルバイト人件費"]+df.loc[:, "賞与社保合計/その他（日払い等）"])

            for index in df.index:
                # 経常利益額がマイナスなら、ゼロに。
                if df.loc[index, "経常利益額"] < 0:
                    df.loc[index, "経常利益額"] = 0

                # 交通費を半分ずつ分散。バイト人件費ゼロなら、全て社員に。
                if df.loc[index, "アルバイト人件費"] == 0:
                    df.loc[index, "社員"] += df.loc[index, "交通費"]
                else:
                    df.loc[index, "社員"] += (df.loc[index, "交通費"] / 2).astype(int)
                    df.loc[index, "アルバイト"] += (df.loc[index, "交通費"] / 2).astype(int)

            df.drop(columns=["社員人件費", "社保合計", "賞与", "アルバイト人件費", "賞与社保合計/その他（日払い等）", "交通費"], inplace=True)

        date_list = list(df_for_chart["日付"])
        sales_list = list(df_for_chart["総売上高"])

    cost_dict = {}
    for index, row in df.T.iterrows():
        cost_dict[index] = list(row)
    del cost_dict["日付"]
    del cost_dict["総売上高"]

    # debug(cost_dict)

    data_dict = {}
    for row in df.itertuples():
        data_dict[row[1].replace('/', ' 年 ') + " 月"] = list(row[2:])

    label_index = list(df.columns[2:])

    context = {
        "date_list": date_list,
        "sales_list": sales_list,
        "data_dict": data_dict,
        "label_index": label_index,
        "cost_dict": cost_dict,
    }
    if flg == "cgy":
        context["store"] = cgy

    return render(request, "display_part/chart.html", context)


def create_df_month(q_list, use: str):
    dflist = []
    df_list_chart = []
    for query in q_list:
        pl_data = PL_data.objects.filter(y_m=query)
        if not pl_data.exists():
            # return render(request, 'display_part/error.html', {"error": "指定された期間が存在しません。"})
            # return HttpResponse(f'{query} 指定された期間が存在しません。')
            return None

        df = pd.read_frame(pl_data, index_col="id")
        # df = df.dropna(how="all", axis=1).dropna(how="any")

        df.set_index("category", drop=False, inplace=True)

        # --年月表示を見やすく--------------
        first_flg = True
        for i, v in df.iterrows():
            if first_flg:  # 始めだけ印をつける
                df.loc[i, "y_m"] = v.y_m[:4] + "/" + v.y_m[4:] + "-"
                first_flg = False
            else:
                df.loc[i, "y_m"] = v.y_m[:4] + "/" + v.y_m[4:]
        # --------------------------------------

        # フードコスト
        costs = FoodCosts.objects.filter(pl_data__in=pl_data)
        df_part = pd.read_frame(costs, index_col="id")
        df_part.set_index("pl_data", inplace=True)  # 店名で揃えてる
        df_part = df_part.assign(フード合計=df_part.iloc[:, :].sum(axis=1))
        df = pandas.concat([df, df_part], axis=1)  # 店名で揃えてる

        # ドリンクコスト
        costs = DrinkCosts.objects.filter(pl_data__in=pl_data)
        df_part = pd.read_frame(costs, index_col="id")
        df_part.set_index("pl_data", inplace=True)  # 店名で揃えてる
        df_part = df_part.assign(ドリンク合計=df_part.iloc[:, :].sum(axis=1))

        df = pandas.concat([df, df_part], axis=1)

        # 仕入れ合計
        df = df.assign(仕入合計=df.loc[:, "フード合計"] + df.loc[:, "ドリンク合計"])

        # 人件費
        costs = LaborCosts.objects.filter(pl_data__in=pl_data)
        df_part = pd.read_frame(costs, index_col="id")
        df_part.set_index("pl_data", inplace=True)  # 店名で揃えてる
        df_part = df_part.assign(人件費合計=df_part.iloc[:, :].sum(axis=1))

        df = pandas.concat([df, df_part], axis=1)

        # Ｆ/Ｌ コスト
        df.insert(len(df.columns), 'Ｆ/Ｌ コスト', df.loc[:, "仕入合計"]+df.loc[:, "人件費合計"])

        # 営業粗利
        df = df.assign(営業粗利=df.loc[:, "amountSold"] - df.loc[:, "Ｆ/Ｌ コスト"])

        # 家賃の列いれかえ
        df = df.assign(家賃=df["rent"])
        df.drop('rent', axis=1, inplace=True)

        # 光熱費・通信費
        costs = UtilityCosts_ComunicationCosts.objects.filter(pl_data__in=pl_data)
        df_part = pd.read_frame(costs, index_col="id")
        df_part.set_index("pl_data", inplace=True)  # 店名で揃えてる
        df_part.insert(len(df_part.columns), "光熱費・通信費　合計", df_part.iloc[:, :].sum(axis=1))  # assignの引数に「光熱費・通信費　合計」が使えないため、insertで。

        df = pandas.concat([df, df_part], axis=1)

        # 広告費
        cost = AdvertisingCosts.objects.filter(pl_data__in=pl_data)
        df_part = pd.read_frame(cost, index_col="id")
        df_part.set_index("pl_data", inplace=True)  # 店名で揃えてる
        df_part.insert(len(df_part.columns), "宣伝・広告費　合計", df_part.iloc[:, :].sum(axis=1))  # assignの引数に「宣伝・広告費　合計」が使えないため、insertで。

        df = pandas.concat([df, df_part], axis=1)

        # その他コスト
        cost = OtherCosts.objects.filter(pl_data__in=pl_data)
        df_part = pd.read_frame(cost, index_col="id")
        df_part.set_index("pl_data", inplace=True)  # 店名で揃えてる
        df_part.insert(len(df_part.columns), "雑費　合計", df_part.iloc[:, :].sum(axis=1))  # assignの引数に「雑費　合計」が使えないため、insertで。

        df = pandas.concat([df, df_part], axis=1)

        # 課税経費　合計
        df.insert(len(df.columns), '課税経費　合計', df.loc[:, "家賃"]+df.loc[:, "光熱費・通信費　合計"] + df.loc[:, "宣伝・広告費　合計"] + df.loc[:, "雑費　合計"])

        # 非課税経費
        cost = TaxExemptExpenses.objects.filter(pl_data__in=pl_data)
        df_part = pd.read_frame(cost, index_col="id")
        df_part.set_index("pl_data", inplace=True)  # 店名で揃えてる
        df_part.insert(len(df_part.columns), "非課税経費　合計", df_part.iloc[:, :].sum(axis=1))

        df = pandas.concat([df, df_part], axis=1)

        # 経費　総計
        df.insert(len(df.columns), '経費　総計', df.loc[:, "課税経費　合計"]+df.loc[:, "非課税経費　合計"])
        # 経常利益額
        df.insert(len(df.columns), '経常利益額', df.loc[:, "amountSold"]-df.loc[:, "Ｆ/Ｌ コスト"]-df.loc[:, "経費　総計"])
        # 支払い消費税（概算）
        if int(query) >= 201910:  # 2019年9月以前は消費税8％
            df.insert(len(df.columns), '支払い消費税（概算）', (df.loc[:, "amountSold"]*100/110*0.1) - ((df.loc[:, "仕入合計"]+df.loc[:, "transportationExpenses"]+df.loc[:, "課税経費　合計"])*100/110*0.1))
        else:
            df.insert(len(df.columns), '支払い消費税（概算）', (df.loc[:, "amountSold"]*100/108*0.08) - ((df.loc[:, "仕入合計"]+df.loc[:, "transportationExpenses"]+df.loc[:, "課税経費　合計"])*100/108*0.08))
        # 税引き後利益
        df.insert(len(df.columns), '税引き後利益', df.loc[:, "経常利益額"]-df.loc[:, "支払い消費税（概算）"])

        df3 = df

        # カラム名を日本語にーーーーー
        for col in range(len(df.columns)):
            for each_dict in my_module.pl_dbdict.values():
                for ver_name, col_name in each_dict.items():
                    if df.columns[col] == ver_name:
                        df.rename(columns={ver_name: col_name}, inplace=True)

        # 家賃きをつけて！！！！！

        # 店順ソート
        df = df.reindex(index=["FES", "Garage あそび", "灯篭",  "罠", "罠一目", "罠中目黒", "reme(小顔矯正)", "Base　あそび",  "楽", "浦安あそび", "よってこ", "本部"])
        df = df.dropna(how="all", axis=1).dropna(how="all")

        dflist.append(df)
        # debug(df.T.iloc[2:, 0:].sum(axis=1)[-1]) # 合計 税引き後利益 確認用

        if use == "chart":
            # 全店1ヶ月分のトータルだけをDF化
            total = df.iloc[0:, 2:].sum()
            total["日付"] = df["日付"][1]  # 日付データを1つ抽出
            total = pandas.DataFrame(total)
            df_list_chart.append(total.T)

    # チャートへのリクエスト時はここでreturn
    if use == "chart":
        return df_list_chart

    # 年をまたぐ際の表示ズレを修正 concatではズレる
    ddd = df.iloc[0:0]
    df = ddd.append(dflist)[df.columns.tolist()]
    # df = pandas.concat(dflist)
    df.fillna(0, inplace=True)

    # 合計insert
    df = df.T
    df.insert(len(df.columns), "total", df.iloc[2:, 0:].sum(axis=1))
    df.iloc[0, len(df.columns)-1] = ""
    df.iloc[1, len(df.columns)-1] = "合計"

    # エクセルへのリクエスト時はここでreturn
    if use == "excel":
        df_for_excel = df.T
        return df_for_excel

    # dfがNanを含むため数値が小数点以下で表示される。それを取り除きたいが文字列も含んでいるためastype(int)が使えない。処理のため一旦別けて再度concatする。
    df_a = df.iloc[:2, :]
    df_b = df.iloc[2:, :].applymap(lambda x: int(Decimal(str(x)).quantize(Decimal('0'), rounding=ROUND_HALF_UP)))
    df = pandas.concat([df_a, df_b])
    # -------------------------------------

    # 売上に対する割合insert
    for i in reversed(range(1, 1+len(df.columns))):
        try:
            df.insert(i, "compa"+str(i), df.iloc[3:, i-1] / df.iloc[2, i-1]*100)
            df.iloc[0, i] = ""
            df.iloc[1, i] = ""
            df.iloc[2, i] = ""
        except ZeroDivisionError as err:
            pass
    df = df.T
    # -------------------------------------

    for w in range(len(df.index)):
        for w2 in range(len(df.columns)):
            if isinstance(df.iloc[w, w2], float):
                df.iloc[w, w2] = f"{df.iloc[w, w2]:.2f}"+"%"
            if type(df.iloc[w, w2]) != str:  # カンマで見やすく
                df.iloc[w, w2] = "{:,}".format(df.iloc[w, w2])

    if use == "table":
        # print(df)
        return df


def create_df_by_paytype(year: str, selected_kinds: str):
    print(year)
    pl_data = PL_data.objects.filter(y_m__icontains=year).order_by("y_m")  # 年で取る
    store_name_list = list(set([s.__str__() for s in list(pl_data)]))

    # モデルと、支払い種類のDB内の名前、を特定
    model, kinds_en = None, ''
    for model, fieldname_dict in my_module.pl_dbdict.items():
        for name_en, name_ja in fieldname_dict.items():
            if name_ja == selected_kinds:
                kinds_en = name_en
                break  # ここでbreakしたら外側のループも抜ける
        else:
            kinds_en = selected_kinds
            continue
        break
    print(model)
    print(kinds_en)

    def create_series(pl_data_by_store, kinds_en: str = "", model=None, is_section=False):
        if kinds_en == "amountSold":
            series = pd.read_frame(pl_data_by_store, index_col="id")["amountSold"]
        elif kinds_en == "rent":
            series = pd.read_frame(pl_data_by_store, index_col="id")["rent"]
        elif is_section:
            costs = model.objects.filter(pl_data__in=pl_data_by_store)
            df_part = pd.read_frame(costs, index_col="id")
            series = df_part.iloc[:, :].sum(axis=1)
        else:
            costs = model.objects.filter(pl_data__in=pl_data_by_store)
            series = pd.read_frame(costs, index_col="id")[kinds_en]
        return series

    dflist = []
    for store in store_name_list:
        pl_data_by_store = pl_data.filter(category=store)
        df_pl = pd.read_frame(pl_data_by_store, index_col="id")[["y_m", "category"]]
        if selected_kinds == "総売上高":  # pl_dataのみの処理
            series = create_series(pl_data_by_store, "amountSold")
            df = pandas.concat([df_pl, series], axis=1)
            df = df.rename(columns={kinds_en: store}).drop("category", axis=1)
        elif selected_kinds == "家賃":  # pl_dataのみの処理
            series = create_series(pl_data_by_store, "rent")
            df = pandas.concat([df_pl, series], axis=1)
            df = df.rename(columns={kinds_en: store}).drop("category", axis=1)

        # 中〜大部門ーーーーーー
        elif selected_kinds in ["フード合計", "ドリンク合計", "仕入合計", "人件費合計", "Ｆ/Ｌ コスト", "営業粗利", "光熱費・通信費　合計", "宣伝・広告費　合計", "雑費　合計", "課税経費　合計", "非課税経費　合計", "経費　総計", "経常利益額", "支払い消費税（概算）", "税引き後利益"]:
            def cheese_model(kinds):
                model = FoodCosts if kinds == "フード合計" else DrinkCosts if kinds == "ドリンク合計" else LaborCosts if kinds == "人件費合計" else UtilityCosts_ComunicationCosts if kinds == "光熱費・通信費　合計" else AdvertisingCosts if kinds == "宣伝・広告費　合計" else OtherCosts if kinds == "雑費　合計" else TaxExemptExpenses
                return model

            def totalize_mediam_kindses(mediam_kinds_list):
                series_list = []
                for mediam_kinds in mediam_kinds_list:
                    model = cheese_model(mediam_kinds)
                    series = create_series(pl_data_by_store, model=model, is_section=True)
                    series_list.append(series)
                df_part = pandas.concat(series_list, axis=1)
                result_series = df_part.iloc[:, :].sum(axis=1)
                result_series = result_series.rename("暫定ごうけい")
                return result_series

            # 大部門。中部門を組み合わせて使用
            if selected_kinds == "仕入合計":
                mediam_kinds_list = ["フード合計", "ドリンク合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)

            elif selected_kinds == "Ｆ/Ｌ コスト":
                mediam_kinds_list = ["フード合計", "ドリンク合計", "人件費合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)

            elif selected_kinds == "営業粗利":
                # Ｆ/Ｌ コスト と一緒ーーーーー
                mediam_kinds_list = ["フード合計", "ドリンク合計", "人件費合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)
                # Ｆ/Ｌ コスト と一緒ーーーーー

                # 他のテーブル(モデル)にあるので。
                amountSold = create_series(pl_data_by_store, "amountSold")

                result_series = amountSold - result_series
                result_series = result_series.rename("暫定ごうけい")

            elif selected_kinds == "課税経費　合計":
                mediam_kinds_list = ["光熱費・通信費　合計", "宣伝・広告費　合計", "雑費　合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)

                # 他のテーブル(モデル)にあるので。
                rent = create_series(pl_data_by_store, "rent")

                result_series = rent + result_series
                result_series = result_series.rename("暫定ごうけい")

            elif selected_kinds == "経費　総計":
                mediam_kinds_list = ["光熱費・通信費　合計", "宣伝・広告費　合計", "雑費　合計", "非課税経費　合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)

                # 他のテーブル(モデル)にあるので。
                rent = create_series(pl_data_by_store, "rent")

                result_series = rent + result_series
                result_series = result_series.rename("暫定ごうけい")

            elif selected_kinds == "経常利益額":
                # 営業粗利 と一緒ーーーーーー
                mediam_kinds_list = ["フード合計", "ドリンク合計", "人件費合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)
                amountSold = create_series(pl_data_by_store, "amountSold")
                result_series = amountSold - result_series

                # 営業粗利 と一緒ーーーーーー

                # 経費　総計 と一緒ーーーーーー
                mediam_kinds_list = ["光熱費・通信費　合計", "宣伝・広告費　合計", "雑費　合計", "非課税経費　合計"]
                result_series2 = totalize_mediam_kindses(mediam_kinds_list)
                rent = create_series(pl_data_by_store, "rent")
                result_series2 = rent + result_series2
                # 経費　総計 と一緒ーーーーーー

                result_series = result_series - result_series2
                result_series = result_series.rename("暫定ごうけい")

            elif selected_kinds == "支払い消費税（概算）":
                df_for_tax = pd.read_frame(pl_data_by_store, index_col="id")[["y_m", "amountSold"]]
                transport_Ex = create_series(pl_data_by_store, kinds_en="transportationExpenses", model=LaborCosts)

                # 仕入合計ーーーーーー
                mediam_kinds_list = ["フード合計", "ドリンク合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)
                # 仕入合計ーーーーーー

                # 課税経費　合計ーーーーーーーーー
                mediam_kinds_list = ["光熱費・通信費　合計", "宣伝・広告費　合計", "雑費　合計"]
                result_series2 = totalize_mediam_kindses(mediam_kinds_list)
                rent = create_series(pl_data_by_store, "rent")
                result_series2 = rent + result_series2
                # 課税経費　合計ーーーーーーーーー

                result_series = result_series.rename("仕入合計")
                result_series2 = result_series2.rename("課税経費　合計")

                df_for_tax = pandas.concat([df_for_tax, transport_Ex, result_series, result_series2], axis=1)

                # y_mを比較できるように、int型に。
                df_for_tax["y_m"] = df_for_tax["y_m"].astype(int)
                # 2019年9月以前は消費税8％
                tax_8_df = df_for_tax[df_for_tax["y_m"] <= 201909]
                tax_8_series = (tax_8_df["amountSold"]*100/108*0.08) - ((tax_8_df["transportationExpenses"] + tax_8_df["仕入合計"] + tax_8_df["課税経費　合計"])*100/108*0.08)
                # 2019年10月以降は消費税10％
                tax_10_df = df_for_tax[df_for_tax["y_m"] >= 201910]
                tax_10_series = (tax_10_df["amountSold"]*100/110*0.1) - ((tax_10_df["transportationExpenses"] + tax_10_df["仕入合計"] + tax_10_df["課税経費　合計"])*100/110*0.1)

                result_series = pandas.concat([tax_8_series, tax_10_series])
                result_series = result_series.rename("暫定ごうけい")

            elif selected_kinds == "税引き後利益":
                # 経常利益額 と一緒ーーーーーー
                mediam_kinds_list = ["フード合計", "ドリンク合計", "人件費合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)
                amountSold = create_series(pl_data_by_store, "amountSold")
                result_series = amountSold - result_series
                mediam_kinds_list = ["光熱費・通信費　合計", "宣伝・広告費　合計", "雑費　合計", "非課税経費　合計"]
                result_series2 = totalize_mediam_kindses(mediam_kinds_list)
                rent = create_series(pl_data_by_store, "rent")
                result_series2 = rent + result_series2
                keijourieki = result_series - result_series2
                # 経常利益額 と一緒ーーーーーー

                # 支払い消費税（概算） と一緒ーーーーーー
                df_for_tax = pd.read_frame(pl_data_by_store, index_col="id")[["y_m", "amountSold"]]
                transport_Ex = create_series(pl_data_by_store, kinds_en="transportationExpenses", model=LaborCosts)
                mediam_kinds_list = ["フード合計", "ドリンク合計"]
                result_series = totalize_mediam_kindses(mediam_kinds_list)
                mediam_kinds_list = ["光熱費・通信費　合計", "宣伝・広告費　合計", "雑費　合計"]
                result_series2 = totalize_mediam_kindses(mediam_kinds_list)
                rent = create_series(pl_data_by_store, "rent")
                result_series2 = rent + result_series2
                result_series = result_series.rename("仕入合計")
                result_series2 = result_series2.rename("課税経費　合計")
                df_for_tax = pandas.concat([df_for_tax, transport_Ex, result_series, result_series2], axis=1)
                df_for_tax["y_m"] = df_for_tax["y_m"].astype(int)
                tax_8_df = df_for_tax[df_for_tax["y_m"] <= 201909]
                tax_8_series = (tax_8_df["amountSold"]*100/108*0.08) - ((tax_8_df["transportationExpenses"] + tax_8_df["仕入合計"] + tax_8_df["課税経費　合計"])*100/108*0.08)
                tax_10_df = df_for_tax[df_for_tax["y_m"] >= 201910]
                tax_10_series = (tax_10_df["amountSold"]*100/110*0.1) - ((tax_10_df["transportationExpenses"] + tax_10_df["仕入合計"] + tax_10_df["課税経費　合計"])*100/110*0.1)
                pay_tax_tatal = pandas.concat([tax_8_series, tax_10_series])
                # 支払い消費税（概算） と一緒ーーーーーー

                result_series = keijourieki - pay_tax_tatal
                result_series = result_series.rename("暫定ごうけい")

            # 中部門
            else:
                mediam_kinds_list = [selected_kinds]
                result_series = totalize_mediam_kindses(mediam_kinds_list)

            df = pandas.concat([df_pl, result_series], axis=1)
            df = df.rename(columns={"暫定ごうけい": store}).drop("category", axis=1)

        # 小部門ーーーーーー
        else:
            series = create_series(pl_data_by_store, kinds_en, model)
            df = pandas.concat([df_pl, series], axis=1)
            df = df.rename(columns={kinds_en: store}).drop("category", axis=1)

        dflist.append(df.set_index("y_m"))

    df_concated = pandas.concat(dflist, axis=1)
    df_concated = df_concated.reindex(columns=["FES", "Garage あそび", "灯篭",  "罠", "罠一目", "罠中目黒", "reme(小顔矯正)", "Base　あそび",  "楽", "浦安あそび", "よってこ", "本部"])
    df_concated = df_concated.dropna(how="all", axis=1)
    df_concated = df_concated.fillna(0)
    # df_concated = df_concated.astype(int)
    df_concated = df_concated.applymap(lambda x: int(Decimal(str(x)).quantize(Decimal('0'), rounding=ROUND_HALF_UP)))  # しっかり四捨五入

    # 年月表示を見やすく
    easy_to_see_date = df_concated.index.map(lambda x: x[:4] + "/" + x[4:])
    dict_to_swap_index = dict(zip(df_concated.index, easy_to_see_date))
    df_concated.rename(index=dict_to_swap_index, inplace=True)

    df_concated.loc["合計"] = df_concated.sum()
    df_concated["合計"] = df_concated.sum(axis=1)

    df_concated = df_concated.T
    df_concated.insert(0, "店名", df_concated.index)

    return df_concated


def create_df_by_store(store, span, use: str):
    pl_data = PL_data.objects.filter(category=store, y_m__icontains=span).order_by("y_m")  # 年で取る
    df2 = pd.read_frame(pl_data, index_col="id")
    # df2 = df2.dropna(how="all", axis=1).dropna(how="any")

    # df2.set_index("category", drop=False, inplace=True)

    # --年月表示を見やすく--------------
    for i, v in df2.iterrows():
        df2.loc[i, "y_m"] = v.y_m[:4] + "/" + v.y_m[4:]
    # --------------------------------------

    costs = FoodCosts.objects.filter(pl_data__in=pl_data)
    df_part = pd.read_frame(costs, index_col="id")
    df_part = df_part.drop('pl_data', axis=1)
    df_part = df_part.assign(フード合計=df_part.iloc[:, :].sum(axis=1))
    df2 = pandas.concat([df2, df_part], axis=1)  # idで揃えてる

    costs = DrinkCosts.objects.filter(pl_data__in=pl_data)
    df_part = pd.read_frame(costs, index_col="id")
    df_part = df_part.drop('pl_data', axis=1)
    df_part = df_part.assign(ドリンク合計=df_part.iloc[:, :].sum(axis=1))
    df2 = pandas.concat([df2, df_part], axis=1)

    # 仕入れ合計
    df2 = df2.assign(仕入合計=df2.loc[:, "フード合計"]+df2.loc[:, "ドリンク合計"])

    costs = LaborCosts.objects.filter(pl_data__in=pl_data)
    df_part = pd.read_frame(costs, index_col="id")
    df_part = df_part.drop('pl_data', axis=1)
    df_part = df_part.assign(人件費合計=df_part.iloc[:, :].sum(axis=1))
    df2 = pandas.concat([df2, df_part], axis=1)

    # Ｆ/Ｌ コスト
    df2.insert(len(df2.columns), 'Ｆ/Ｌ コスト', df2.loc[:, "仕入合計"]+df2.loc[:, "人件費合計"])

    # 営業粗利
    df2 = df2.assign(営業粗利=df2.loc[:, "amountSold"]+df2.loc[:, "Ｆ/Ｌ コスト"])

    # 家賃の列いれかえ
    df2 = df2.assign(家賃=df2["rent"])
    df2.drop('rent', axis=1, inplace=True)

    costs = UtilityCosts_ComunicationCosts.objects.filter(pl_data__in=pl_data)
    df_part = pd.read_frame(costs, index_col="id")
    df_part = df_part.drop('pl_data', axis=1)
    df_part.insert(len(df_part.columns), "光熱費・通信費　合計", df_part.iloc[:, :].sum(axis=1))
    df2 = pandas.concat([df2, df_part], axis=1)

    costs = AdvertisingCosts.objects.filter(pl_data__in=pl_data)
    df_part = pd.read_frame(costs, index_col="id")
    df_part = df_part.drop('pl_data', axis=1)
    df_part.insert(len(df_part.columns), "宣伝・広告費　合計", df_part.iloc[:, :].sum(axis=1))
    df2 = pandas.concat([df2, df_part], axis=1)

    costs = OtherCosts.objects.filter(pl_data__in=pl_data)
    df_part = pd.read_frame(costs, index_col="id")
    df_part = df_part.drop('pl_data', axis=1)
    df_part.insert(len(df_part.columns), "雑費　合計", df_part.iloc[:, :].sum(axis=1))
    df2 = pandas.concat([df2, df_part], axis=1)

    # 課税経費　合計
    df2.insert(len(df2.columns), '課税経費　合計', df2.loc[:, "家賃"]+df2.loc[:, "光熱費・通信費　合計"] + df2.loc[:, "宣伝・広告費　合計"] + df2.loc[:, "雑費　合計"])

    costs = TaxExemptExpenses.objects.filter(pl_data__in=pl_data)
    df_part = pd.read_frame(costs, index_col="id")
    df_part = df_part.drop('pl_data', axis=1)
    df_part.insert(len(df_part.columns), "非課税経費　合計", df_part.iloc[:, :].sum(axis=1))
    df2 = pandas.concat([df2, df_part], axis=1)

    # 経費　総計
    df2.insert(len(df2.columns), '経費　総計', df2.loc[:, "課税経費　合計"]+df2.loc[:, "非課税経費　合計"])
    # 経常利益額
    df2.insert(len(df2.columns), '経常利益額', df2.loc[:, "amountSold"]-df2.loc[:, "Ｆ/Ｌ コスト"]-df2.loc[:, "経費　総計"])

    # 支払い消費税（概算）
    for i, v in df2.iterrows():
        if int(v["y_m"].replace("/", "")) >= 201909:
            df2.loc[i, "支払い消費税（概算）"] = ((df2.loc[i, "amountSold"]*100/110*0.1) - ((df2.loc[i, "仕入合計"]+df2.loc[i, "transportationExpenses"]+df2.loc[i, "課税経費　合計"])*100/110*0.1))
        else:
            df2.loc[i, "支払い消費税（概算）"] = ((df2.loc[i, "amountSold"]*100/108*0.08) - ((df2.loc[i, "仕入合計"]+df2.loc[i, "transportationExpenses"]+df2.loc[i, "課税経費　合計"])*100/108*0.08))

    # 税引き後利益
    df2.insert(len(df2.columns), '税引き後利益', df2.loc[:, "経常利益額"]-df2.loc[:, "支払い消費税（概算）"])

    df2 = df2.dropna(how="all", axis=1).dropna(how="all")
    df2.fillna(0, inplace=True)

    # カラム名を日本語にーーーーー
    for col in range(len(df2.columns)):
        for each_dict in my_module.pl_dbdict.values():
            for ver_name, col_name in each_dict.items():
                if df2.columns[col] == ver_name:
                    df2.rename(columns={ver_name: col_name}, inplace=True)

    # チャートへのリクエスト時はここでreturn
    if use == "chart":
        return df2

    # 合計insertーーーーーーー
    df2 = df2.T
    df2.insert(len(df2.columns), "total", df2.iloc[2:, 0:].sum(axis=1))
    # print(df2)
    df2.iloc[0, len(df2.columns)-1] = ""
    df2.iloc[1, len(df2.columns)-1] = "合計"

    # df2.drop(df2.index[df2["total"] == 0],inplace=True) # allゼロの行を消す

    # エクセルへのリクエスト時はここでreturn
    if use == "excel":
        return df2.T

    # dfがNanを含むため数値が小数点以下で表示される。それを取り除きたいが文字列も含んでいるためastype(int)が使えない。処理のため一旦別けて再度concatする。
    df_a = df2.iloc[:2, :]
    df_b = df2.iloc[2:, :].applymap(lambda x: int(Decimal(str(x)).quantize(Decimal('0'), rounding=ROUND_HALF_UP)))
    df2 = pandas.concat([df_a, df_b])
    # -------------------------------------

    # 売上に対する割合insert
    for i in reversed(range(1, 1+len(df2.columns))):
        try:
            df2.insert(i, "compa"+str(i), df2.iloc[3:, i-1] / df2.iloc[2, i-1]*100)
            df2.iloc[0, i] = ""
            df2.iloc[1, i] = ""
            df2.iloc[2, i] = ""
        except ZeroDivisionError as err:
            pass
    df2 = df2.T
    # -------------------------------------

    for w in range(len(df2.index)):
        for w2 in range(len(df2.columns)):
            if isinstance(df2.iloc[w, w2], float):
                df2.iloc[w, w2] = f"{df2.iloc[w, w2]:.2f}"+"%"
            if type(df2.iloc[w, w2]) != str:
                df2.iloc[w, w2] = "{:,}".format(df2.iloc[w, w2])
    if use == "table":
        # print(df2)
        return df2, pl_data
