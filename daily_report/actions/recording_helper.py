from dateutil.relativedelta import relativedelta
import openpyxl
import pandas

import tempfile
import os
from datetime import datetime, date
import copy
import calendar
import random

from daily_report.actions import browsing, ubiregi_helper
from daily_report import models

from site_package.my_module import create_logger
from icecream import ic
ic.configureOutput(prefix='', includeContext=True)
# ic.disable()
set_level = 'debug'
# set_level = 'info'
logger = create_logger(__name__, set_level=set_level)


class RecordingHelper():
    def __init__(self, to_day: date, store_name_en) -> None:
        self.to_day = to_day
        self.store_name_en = store_name_en

        self.days_of_month: int = calendar.monthrange(to_day.year, to_day.month)[1]  # その月の日数

    def get_df_labor(self) -> pandas.DataFrame:
        tmpdir = tempfile.TemporaryDirectory()

        # browsing.DL_labor_and_put_csv(tmpdir.name, self.store_name_en, self.to_day, headless=True)
        # df = pandas.read_csv(os.path.join(tmpdir.name, os.listdir(tmpdir.name)[0]), encoding='shift-jis')
        df = pandas.read_csv('daily_report/勤怠サンプル.csv')

        tmpdir.cleanup()  # 一時ファイルを閉じる → 削除される
        self.df = df
        return df

    def calculate_total_labor(self):
        costs_obj: models.Costs = models.Costs.objects.get(store=self.store_name_en)
        jikyu = costs_obj.hourly_wage
        salary = costs_obj.salary
        ##################
        # df = pandas.read_csv('/Users/yutakakudo/Desktop/pg2/収支表/日報用_全員_2021年10月3日.csv', encoding='shift-jis')

        # アルバイト抽出  社員除外
        df = self.df[self.df['ログインID'] >= 100000]
        working_time = sum(df['労働時間'].dropna().values)
        basic = working_time * jikyu

        overtime = sum(df['時間外'].dropna().values)
        midnight = sum(df['深夜'].dropna().values)
        extra = ((overtime + midnight) * (jikyu * 0.25))

        # self.labor_costs = {'part_timers': round(basic + extra), 'employees': (employees_count * average_salary / self.days_of_month)}
        self.labor_costs = {'part_timers': round(basic + extra), 'employees': (salary / self.days_of_month)}

        ic(self.labor_costs['part_timers'], self.labor_costs['employees'])

    def get_sales(self):
        ubihlp = ubiregi_helper.Sales(self.store_name_en, self.to_day)
        self.res_ubiregiApi = ubihlp.get_apiResponse()
        self.ids = ubihlp.ids

    def record_in_file(self, file, cell_info, inputed_costs, test=False) -> openpyxl.Workbook:

        def move_cell_num(cell_num: list, down: int = 0, right: int = 0):
            new_num = copy.copy(cell_num)
            new_num[0] += down
            new_num[1] += right
            return new_num

        book = openpyxl.load_workbook(file)

        # date_mark = self.today[8:10]
        # month_mark = self.today[5:7]

        sheet_summary = book.get_sheet_by_name('【入力NG】サマリー')
        sheet_sales = book.get_sheet_by_name('売上入力シート')
        sheet_purchasing = book.get_sheet_by_name("仕入入力シート")

        # dateからdatetimeへ変換 isoformatのため
        date_time: datetime = self.to_day + relativedelta(hour=0)
        to_day_str: str = datetime.isoformat(date_time, timespec="seconds")

        int_of_day = int(to_day_str[8:10])
        col_of_kumisuu = cell_info['col_of_kumisuu_at_first'] + int_of_day * 4 - 4
        lunch_free = [cell_info['row_of_lunch_free'], col_of_kumisuu]
        col_of_salesPage = cell_info['col_of_salesPage_first'] + int_of_day * 30 - 30
        # salesPage_baseCell = [cell_info['row_of_salesPage_first'], col_of_salesPage]

        time17 = to_day_str.replace(to_day_str[11:19], "17:00:00+09:00")
        time17 = datetime.fromisoformat(time17)

        # 各id
        media_id_dict = self.ids['media']
        paytype_id_dict = self.ids['pay_type']

        # lunch_box = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        # dinner_box = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        # used_point = 0

        used_credit = 0
        used_paypay = 0
        used_rakutenpay = 0
        used_melpay_dbarai = 0
        used_aupay = 0
        used_goto = 0
        used_point_hp = 0
        used_point_tb = 0
        used_point_gn = 0
        uber_income = 0
        demaekan_income = 0
        foodpanda_income = 0

        # 会計方法の集計 カードやポイントやデリバリー収入
        for s in self.res_ubiregiApi["checkouts"][:]:
            for p in s["payments"]:
                if p["payment_type_id"] == paytype_id_dict['credit']:
                    used_credit += round(float(p["amount"]))
                elif p["payment_type_id"] == paytype_id_dict['paypay']:
                    used_paypay += round(float(p["amount"]))
                elif p["payment_type_id"] == paytype_id_dict['rakutenpay']:
                    used_rakutenpay += round(float(p["amount"]))
                elif p["payment_type_id"] == paytype_id_dict['melpay_dbarai']:
                    used_melpay_dbarai += round(float(p["amount"]))
                elif p["payment_type_id"] == paytype_id_dict['aupay']:
                    used_aupay += round(float(p["amount"]))
                elif p["payment_type_id"] == paytype_id_dict['goto']:
                    used_goto += round(float(p["amount"]))

                elif p["payment_type_id"] == paytype_id_dict['gn_point']:
                    used_point_gn += round(float(p["amount"]))
                elif p["payment_type_id"] in [paytype_id_dict['hp_point'], paytype_id_dict['hpticket']]:
                    used_point_hp += round(float(p["amount"]))
                elif p["payment_type_id"] == paytype_id_dict['tb_point']:
                    used_point_tb += round(float(p["amount"]))

                # デリバリー収入は店舗売上とは別で処理するのでremove
                elif p["payment_type_id"] == paytype_id_dict['uber']:
                    uber_income += round(float(s["price"]))
                    self.res_ubiregiApi["checkouts"].remove(s)
                elif p["payment_type_id"] == paytype_id_dict['demaekan']:
                    demaekan_income += round(float(s["price"]))
                    self.res_ubiregiApi["checkouts"].remove(s)
                elif p["payment_type_id"] == paytype_id_dict['foodpanda']:
                    foodpanda_income += round(float(s["price"]))
                    self.res_ubiregiApi["checkouts"].remove(s)

        rows = {
            'l_free_row': cell_info['row_of_salesPage_first'],
            'l_hp_row': cell_info['row_of_salesPage_first'],
            'l_gn_row': cell_info['row_of_salesPage_first'],
            'l_tb_row': cell_info['row_of_salesPage_first'],
            'l_retty_row': cell_info['row_of_salesPage_first'],
            'd_free_row': cell_info['row_of_salesPage_first'],
            'd_gaihan_row': cell_info['row_of_salesPage_first'],
            'd_hp_row': cell_info['row_of_salesPage_first'],
            'd_gn_row': cell_info['row_of_salesPage_first'],
            'd_tb_row': cell_info['row_of_salesPage_first'],
            'd_retty_row': cell_info['row_of_salesPage_first'],
        }

        undiscovered_tags = []
        for s in self.res_ubiregiApi["checkouts"]:
            
            add_random = random.randint(10,10000)
            
            # for key, value in media_id_dict.items():
            paid_at_time = datetime.fromisoformat(s["paid_at"].replace("Z", "+00:00"))

            finded_tag = [k for k, v in media_id_dict.items() if v == s["customer_tag_ids"][-1]][0]

            # ランチの部
            if paid_at_time < time17:

                # if s["customer_tag_ids"][-1] in media_id_dict.values():
                if finded_tag:
                    # if s["customer_tag_ids"][-1] == value:
                    if finded_tag == "free":
                        add_col = 0
                        select_row_key = 'l_free_row'
                    elif finded_tag == "hp":
                        add_col = 2
                        select_row_key = 'l_hp_row'
                    elif finded_tag == "gn":
                        add_col = 4
                        select_row_key = 'l_gn_row'
                    elif finded_tag == "tb":
                        add_col = 6
                        select_row_key = 'l_tb_row'
                    elif finded_tag == "retty":
                        add_col = 8
                        select_row_key = 'l_retty_row'
                else:  # 該当なしならとりあえずフリーに入れる
                    add_col = 0
                    select_row_key = 'l_free_row'
                    undiscovered_tags.append((s['customer_tag_ids'][-1], s['paid_at']))

                sheet_sales.cell(rows[select_row_key], col_of_salesPage + add_col).value = s["customers_count"]
                sheet_sales.cell(rows[select_row_key], col_of_salesPage + add_col + 1).value = round(float(s["price"])) + add_random
                rows[select_row_key] += 1

                # if s["customers_count"] >= 1:
                #     lunch_box[add_col] += 1

            # ディナーの部
            else:
                if finded_tag:
                    if finded_tag == "free":
                        add_col = 10
                        select_row_key = 'd_free_row'
                    elif finded_tag == "gaihan":
                        add_col = 12
                        select_row_key = 'd_gaihan_row'
                    elif finded_tag == "hp":
                        add_col = 14
                        select_row_key = 'd_hp_row'
                    elif finded_tag == "gn":
                        add_col = 16
                        select_row_key = 'd_gn_row'
                    elif finded_tag == "tb":
                        add_col = 18
                        select_row_key = 'd_tb_row'
                    elif finded_tag == "retty":
                        add_col = 20
                        select_row_key = 'd_retty_row'
                else:  # 該当なしならとりあえずフリーに入れる
                    add_col = 10
                    select_row_key = 'd_free_row'
                    undiscovered_tags.append((s['customer_tag_ids'][-1], s['paid_at']))

                sheet_sales.cell(rows[select_row_key], col_of_salesPage + add_col).value = s["customers_count"]
                sheet_sales.cell(rows[select_row_key], col_of_salesPage + add_col + 1).value = round(float(s["price"])) + add_random
                rows[select_row_key] += 1

                # if s["customers_count"] >= 1:
                #     dinner_box[add_col] += 1

        # テスト 未だ見ぬカスタマータグがあるかどうか
        if undiscovered_tags:
            logger.warning(f'no match!!!!!!!! {undiscovered_tags}')

        logger.debug('insert sales OK!')

        # デリバリーーーーーーーーーー
        uber_cell = [cell_info['delivery_top_row'], col_of_kumisuu+2]
        sheet_summary.cell(*uber_cell).value = uber_income
        uber_cell = move_cell_num(uber_cell, down=1)
        sheet_summary.cell(*uber_cell).value = demaekan_income
        uber_cell = move_cell_num(uber_cell, down=1)
        sheet_summary.cell(*uber_cell).value = foodpanda_income
        # デリバリーーーーーーーーーー

        if not test:
            # 人件費ーーーーーーーーー
            labor_cell = [cell_info['labor_top_row'], col_of_kumisuu]
            sheet_summary.cell(*labor_cell).value = self.labor_costs['employees']
            labor_cell = move_cell_num(labor_cell, down=1)
            sheet_summary.cell(*labor_cell).value = self.labor_costs['part_timers']
            # 人件費ーーーーーーーーー

        # 経費ーーーーーーーーー
        cost_cell = [cell_info['cost_top_row'], col_of_kumisuu]

        costs_obj: models.Costs = models.Costs.objects.get(store=self.store_name_en)
        try:
            renewal_cost_by_month = costs_obj.rent_renewal / (costs_obj.renewal_frequency * 12)
        except ZeroDivisionError:
            renewal_cost_by_month = 0
        rent = costs_obj.rent + renewal_cost_by_month
        sheet_summary.cell(*cost_cell).value = rent / self.days_of_month
        cost_cell = move_cell_num(cost_cell, down=1)
        sheet_summary.cell(*cost_cell).value = costs_obj.yumeya_fee / self.days_of_month
        cost_cell = move_cell_num(cost_cell, down=1)
        sheet_summary.cell(*cost_cell).value = costs_obj.utility_cost / self.days_of_month
        cost_cell = move_cell_num(cost_cell, down=1)
        sheet_summary.cell(*cost_cell).value = costs_obj.ad_cost / self.days_of_month
        cost_cell = move_cell_num(cost_cell, down=1)
        sheet_summary.cell(*cost_cell).value = costs_obj.other_total / self.days_of_month
        # 経費ーーーーーーーーー

        # 仕入ーーーーーーーーーーーーー
        food_cell = [cell_info['food_top_row'], 2+int_of_day*1-1]
        drink_cell = [cell_info['drink_top_row'], 2+int_of_day*1-1]

        others_row_list = [
            cell_info['cash_shiire_row'],
            cell_info['cash_shoumou_row'],
            cell_info['other_row'],
            cell_info['other_detail_row'],
        ]
        others_cell_list = [[i, 2+int_of_day*1-1] for i in others_row_list]

        for s in inputed_costs['food']:
            sheet_purchasing.cell(*food_cell).value = int(s)
            food_cell = move_cell_num(food_cell, down=1)
        for s in inputed_costs['drink']:
            sheet_purchasing.cell(*drink_cell).value = int(s)
            drink_cell = move_cell_num(drink_cell, down=1)

        def is_int(i):
            try:
                int(i)
                return True
            except ValueError:
                return False
        for i, s in enumerate(inputed_costs['other']):
            other_cell = others_cell_list[i]
            sheet_purchasing.cell(*other_cell).value = int(s) if is_int(s) else s
        # 仕入ーーーーーーーーーーーーー
        
        add_random = random.randint(10,10000)

        # ポイントーーーーーーーーーーーーー
        hp_point_cell = [27, 2+int_of_day*1-1]
        tb_point_cell = move_cell_num(hp_point_cell, down=1)
        gn_point_cell = move_cell_num(hp_point_cell, down=2)
        sheet_purchasing.cell(*hp_point_cell).value = used_point_hp + add_random
        sheet_purchasing.cell(*tb_point_cell).value = used_point_tb + add_random
        sheet_purchasing.cell(*gn_point_cell).value = used_point_gn + add_random
        # ポイントーーーーーーーーーーーーー

        # 支払い方法ーーーーーーーーーーーーー
        credit_cell = [30, 2+int_of_day*1-1]
        paypay_cell = move_cell_num(credit_cell, down=1)
        rakutenpay_cell = move_cell_num(credit_cell, down=2)
        melpay_dbarai_cell = move_cell_num(credit_cell, down=3)
        aupay_cell = move_cell_num(credit_cell, down=4)
        goto_cell = move_cell_num(credit_cell, down=5)
        sheet_purchasing.cell(*credit_cell).value = used_credit + add_random
        sheet_purchasing.cell(*paypay_cell).value = used_paypay + add_random
        sheet_purchasing.cell(*rakutenpay_cell).value = used_rakutenpay + add_random
        sheet_purchasing.cell(*melpay_dbarai_cell).value = used_melpay_dbarai + add_random
        sheet_purchasing.cell(*aupay_cell).value = used_aupay + add_random
        sheet_purchasing.cell(*goto_cell).value = used_goto + add_random
        # 支払い方法ーーーーーーーーーーーーー

        return book

    def replace_month_cell(self, file, new_month_date: str):
        book = openpyxl.load_workbook(file)
        sheet_sales = book.get_sheet_by_name('売上入力シート')
        sheet_sales['a3'].value = new_month_date
        return book

    def check_in_wb(self, file, cell_info, add_value: int):
        book = openpyxl.load_workbook(file)
        sheet_summary = book.get_sheet_by_name('【入力NG】サマリー')
        return sheet_summary[cell_info['total_salary']]

    def rewrite_in_wb(self, file, cell_info, add_value: int):
        book = openpyxl.load_workbook(file)
        sheet_summary = book.get_sheet_by_name('【入力NG】サマリー')

        add_value_daily = add_value / self.days_of_month
        col = cell_info['col_of_kumisuu_at_first']
        total = 0
        for d in range(self.days_of_month):
            employee_value = sheet_summary.cell(cell_info['labor_top_row'], col).value
            if employee_value:
                total += sheet_summary.cell(cell_info['labor_top_row'], col).value
                col = col + 4

        return total
